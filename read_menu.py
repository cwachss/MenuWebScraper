import openpyxl as xl


def read_menu(menu, menu_number):
    menu_workbook = xl.load_workbook(menu)
    menu_sheet = menu_workbook[menu_number]
    menu_items = []
    for row in range(2, menu_sheet.max_row + 1):
        ingredient = menu_sheet.cell(row, 1).value
        menu_items.append(ingredient)

    return menu_items
